#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Update Excel sheets by inserting only brand-new rows while keeping styles."""

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# ─────────────────────────────────────
# Excel 読み取り（変更なし）
# ─────────────────────────────────────
def read_sheet(path: str, sheet: str) -> pd.DataFrame:
    if os.path.exists(path):
        try:
            return pd.read_excel(path, sheet_name=sheet)
        except ValueError:
            pass
    return pd.DataFrame()


# ─────────────────────────────────────
# ★ 重複判定を「リンク」列のみに変更
# ─────────────────────────────────────
def diff_by_row(df_new: pd.DataFrame, df_old: pd.DataFrame) -> pd.DataFrame:
    """
    旧シートの「リンク」列と一致する行を df_new から除外して返す。
    """
    # 新データ側の完全重複を削除（必要に応じて）
    df_new = df_new.drop_duplicates().copy()

    if df_old.empty or 'リンク' not in df_new.columns:
        return df_new

    # 旧シートのリンク集合
    old_links = set(df_old['リンク'].astype(str).str.strip())

    # 新データのリンクでフィルタ
    df_new = df_new[~df_new['リンク'].astype(str).str.strip().isin(old_links)]
    return df_new


# ─────────────────────────────────────
def prepend_rows_keep_style(path: str, sheet: str, df_new: pd.DataFrame) -> None:
    wb = load_workbook(path) if os.path.exists(path) else Workbook()
    ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
    df_old = read_sheet(path, sheet) if ws.max_row >= 2 else pd.DataFrame()

    df_diff = diff_by_row(df_new, df_old)
    if df_diff.empty:
        wb.save(path); wb.close(); return

    if ws.max_row == 0:
        ws.append(list(df_diff.columns))          # ヘッダー

    ws.insert_rows(2, len(df_diff))               # 追加行ぶん空ける
    for r, row in enumerate(dataframe_to_rows(df_diff, index=False, header=False), 2):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)

    wb.save(path); wb.close()


# ─────────────────────────────────────
def main() -> None:
    output_file = 'new_releases.xlsx'

    from scrapers.asahi    import fetch as fetch_asahi
    from scrapers.cocacola import fetch as fetch_cocacola
    from scrapers.kirin    import fetch as fetch_kirin
    from scrapers.suntory  import fetch as fetch_suntory

    sources = {
        'Asahi':    fetch_asahi(),
        'CocaCola': fetch_cocacola(),
        'Kirin':    fetch_kirin(),
        'Suntory':  fetch_suntory(),
    }

    for sheet, df in sources.items():
        if not df.empty:
            prepend_rows_keep_style(output_file, sheet, df)

    print(f'Updated → {output_file}')


if __name__ == '__main__':
    main()